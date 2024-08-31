"""Task for Data Engineer role at BFI

Main Purpose: In this task, we will work with the latest UK box office data from July and August 2024. 
The objective is to collate data from the top 15 movies from each BFI box office Excel file. 
We will then generate individual sheets for the top 15 movies in the file bfi-weekend-box-office-report-2024-08-23-25.xls, 
containing information about the movies and their weekend box office income.

Steps:
#1 Extraction: Extract all data from the box office Excel files.
#2 Selection: Identify the top 15 movie titles for the weekend of 23rd-25th August 2024.
#3 API Retrieval: Find the OMDB API movie page for each of the 15 movies.
#4 Sheet Generation: Create individual sheets in a new file for each movie, containing data retrieved from OMDB.
#5 Data Integration: Paste the weekend box office data into each corresponding sheet.

"""

#Importing the required library
import xlrd
import os
import re
import requests
import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import Font
#xlrd is a free library that allows us to open and edit .XLS files 
#os is a library that allows us to call the API_KEY for OMDB without sharing it on Github
#re is a library that allows us to call commands for search patters
#requests to handle API calls
#pandas to work with dataframe
#dotenv is to call the API KEY stored in a hidden file

#importing api_key to connect to the OMDB API
load_dotenv()
api_key = os.getenv("API_KEY")



#Custom functions

#1 Extraction: Extract all data from the box office Excel files.
def create_movie_database():
    folder_path = 'data'
    #creating an empty database to store all the data
    box_office_database = {}
    for filename in os.listdir(folder_path):
        #If condition to see if the file in in the .xls format
        if filename.endswith(".xls"):
            # use function to retrieve data from each box office file
            workbook = xlrd.open_workbook("data/"+filename)
            # Select sheet containing box office data
            box_office_sheet = workbook.sheet_by_index(0)
            #saving the date from the filename into a separate variable
            box_office_date = 0
            pattern = r"(\d{4}-\d{2}-\d{2}-\d{2})"
            match = re.search(pattern, filename)
            #saving the date stored in the filename
            box_office_date = match.group(1)
            
            for i in range(2,18):
                # If conditions check each cell location in row 1 and see if it's a number or not  
                if isinstance(box_office_sheet.cell_value(rowx=i, colx=0), (int, float)):
                    #saving each row for that movie into separate variables
                    rank = int(box_office_sheet.cell_value(rowx=i, colx=0))
                    movie_title = box_office_sheet.cell_value(rowx=i, colx=1)
                    country = box_office_sheet.cell_value(rowx=i, colx=2)
                    weekend_gross = box_office_sheet.cell_value(rowx=i, colx=3)
                    distributor = box_office_sheet.cell_value(rowx=i, colx=4)
                    change_last_week = box_office_sheet.cell_value(rowx=i, colx=5)
                    weeks_release = int(box_office_sheet.cell_value(rowx=i, colx=6))
                    numbers_of_cinema = int(box_office_sheet.cell_value(rowx=i, colx=7))
                    site_average = box_office_sheet.cell_value(rowx=i, colx=8)
                    total_gross_to_date = box_office_sheet.cell_value(rowx=i, colx=9)
                    #crearting a unique key for each entrance in the dictionary and saving all the data
                    key = movie_title + "_" + box_office_date
                    #adding data to the dictionary
                    box_office_database[key] = (rank,movie_title,country,weekend_gross,distributor,change_last_week,weeks_release,numbers_of_cinema,site_average,total_gross_to_date,box_office_date)
    return box_office_database


#Convert a title movie, removing spaces and anniversary string
def convert_movie_title_name(movie_title):
    #check if the movie title contains anniversary string in parentheses
    pattern = r"\(.*Anniversary.*\)"
    if re.search(pattern, movie_title, re.IGNORECASE): 
        #if yes, remove anything contained in the parentheses and remove trailing characters  
        movie_title = re.sub(pattern, "", movie_title).rstrip()
    #remove spaces and replace them with + 
    movie_title = movie_title.replace(" ", "+")
    return movie_title


#3 API Retrieval: Find the OMDB API movie page for each of the 15 movies.
def find_omdb_details(key,movie_title):
    url = "http://www.omdbapi.com/" 
    pattern = r"\(.*Anniversary.*\)"
    if re.search(pattern, key, re.IGNORECASE): 
        #if yes, it means it's an anniversary movie and we won't use the 2024 year parameter
        params = {
            'apikey': api_key,
            't': movie_title, 
        }
    else:
        params = {
            'apikey': api_key,
            't': movie_title, 
            'y': 2024
        }
    #Expeption for this movie which is an old release but not an anniversary
    if (movie_title == 'The+Amazing+Spider-Man'):
        params = {
            'apikey': api_key,
            't': movie_title, 
        } 
    response = requests.get(url, params=params)
    movie_details = response.json()    
    #returning the API call
    return movie_details




#1 Extraction: Extract all data from the box office Excel files.

movie_dataset = create_movie_database()


#2 Selection: Identify the top 15 movie titles for the weekend of 23rd-25th August 2024.

# Load the workbook used for this task
workbook = xlrd.open_workbook('data/bfi-weekend-box-office-report-2024-08-23-25.xls')
# Select sheet containing box office data
box_office_sheet = workbook.sheet_by_index(0)
# Store the movie_title, rank and position into a dictionary
movie_list = {}
# Creating a loop that goes from number 2 to 18
for i in range(2,18):
    # If conditions check each cell location in row 1 and see if it's a number or not  
    if isinstance(box_office_sheet.cell_value(rowx=i, colx=0), (int, float)):
        #if condition is satisfied we are saving the movie title into a variable
        movie_title = box_office_sheet.cell_value(rowx=i, colx=1)
        #calling the custom function to format the name
        movie_title_formatted = convert_movie_title_name(movie_title)
        #saving the original movie title as a key, and its movie title formatted
        movie_list[movie_title] = (movie_title_formatted)



#4 Sheet Generation: Create individual sheets in a new file for each movie, containing data retrieved from OMDB.

#creating a new excel file to where store the output data
with pd.ExcelWriter('Task_output.xlsx', engine='openpyxl') as writer:
    #going through each movie in the list saved previously
    for key, values in movie_list.items():
        #retrieving the OMDB data stored in their API
        movie_details = find_omdb_details(key, values)

        #4: Generate Individual sheets for each movie in a new file and store each information     
        #converting the movie_details data into a pandas dataframe
        df = pd.DataFrame([movie_details])
        #Transposing the Data Frame for a better look
        df = df.T

        #cleaning the movie title to avoid error in naming the sheet
        key_clean = re.sub(r'[\\/?:*|"<>]', '', key) 
        key_clean = key_clean[:31]
        #adding movie details information to the relevant sheet
        df.to_excel(writer, sheet_name=key_clean, index=True,header=False)



        #5 Data Integration: Paste the weekend box office data into each corresponding sheet.
         
        #calling the same sheet to add the box office performance
        worksheet = writer.sheets[key_clean]
        #finding the next free line
        next_free_row = worksheet.max_row + 2 
        #adding the title section
        title_section = worksheet.cell(row=next_free_row, column=1, value="Weekend Box Office Performance")
        title_section.font = Font(bold=True)
        next_free_row += 1
        #adding the column name
        worksheet.cell(row=next_free_row, column=1, value="Weeks on release")
        worksheet.cell(row=next_free_row, column=2, value="Weekend Date Range")
        worksheet.cell(row=next_free_row, column=3, value="Weekend Gross")
        worksheet.cell(row=next_free_row, column=4, value="Percentage change on last week")
        worksheet.cell(row=next_free_row, column=5, value="Number of screens")
        worksheet.cell(row=next_free_row, column=6, value="Site Average")
        worksheet.cell(row=next_free_row, column=7, value="Total Gross to date")
        next_free_row += 1
        #empty list to store the box office information of each single movie
        box_office_data = []
        #finding the relevant rows from the movie_database
        for keys,values in movie_dataset.items():
            #removing the date from the key to compare the movie titles
            trimmed_key = keys.split('_')[0]
            #checking we are on the right row for that movie
            if trimmed_key == key:
                current_week = values[6]
                weekend = keys.split('_')[1]
                weekend_gross = values[3]
                change_of_last_week = values[5]
                number_of_cinema = values[7]
                site_average = values[8]
                total_gross_to_date = values[9]
                #add these details to the list created before
                box_office_data.append((current_week, weekend, weekend_gross, change_of_last_week, number_of_cinema, site_average, total_gross_to_date))

        #sorting the box office list
        box_office_data_sorted = sorted(box_office_data, key=lambda x: x[1], reverse=False)

        #adding to individual file the box office information
        for data in box_office_data_sorted:
            current_week, weekend, weekend_gross, change_of_last_week, number_of_cinema, site_average, total_gross_to_date = data
            worksheet.cell(row=next_free_row, column=1, value=current_week)  # Column A
            worksheet.cell(row=next_free_row, column=2, value=weekend)  # Column A
            worksheet.cell(row=next_free_row, column=3, value=weekend_gross)  # Column B
            worksheet.cell(row=next_free_row, column=4, value=change_of_last_week)  # Column C
            worksheet.cell(row=next_free_row, column=5, value=number_of_cinema)  # Column D
            worksheet.cell(row=next_free_row, column=6, value=site_average)  # Column E
            worksheet.cell(row=next_free_row, column=7, value=total_gross_to_date)  # Column F
            next_free_row += 1

print("Task Output file created successfully.")








